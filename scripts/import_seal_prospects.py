"""
Clean & normalise the marketing SEAL prospects workbook into a tidy JSON the
backend imports. Mirrors scripts/import_prospects.py (FET), but handles this
workbook's own quirks: two tables stacked in one sheet (BODA BODAS holds a
second "ASSOCIATION ( BIKERS )" table below a blank row), a duplicated
HOSPITALS sheet, "not publicaly listed" placeholders, non-breaking spaces,
phones stored as bare integers, and multi-value email/phone cells.

Run:  python3 scripts/import_seal_prospects.py
Out:  backend/database/data/seal-prospects.json
Then: php artisan prospects:import --product=SEAL   (in backend/)
"""

import os, re, json
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "SEAL PROSPECTS SEAL.xlsx")
OUT_DIR = os.path.join(ROOT, "backend", "database", "data")
OUT = os.path.join(OUT_DIR, "seal-prospects.json")
os.makedirs(OUT_DIR, exist_ok=True)

PRODUCT = "SEAL"
SOURCE = "marketing import 2026-07 (SEAL)"

# Sheet title (stripped, lowercased) -> canonical category.
CATEGORY = {
    "phamarcies": "PHARMACY",
    "hospitals": "HOSPITAL",
    "hospitals (2)": "HOSPITAL",          # exact duplicate sheet; deduped by key below
    "first responders": "FIRST_RESPONDER",
    "national sports associations": "SPORTS_ASSOCIATION",
    "mines and quaries": "MINING_QUARRY",
    "manufacturing industries": "MANUFACTURING",
    "tavel co.s": "TRAVEL_COMPANY",
    "boda bodas": "BODA_BODA",
}

# A sheet can stack a second table under a blank row with its own header. Map the
# header's first cell (lowercased, punctuation-loose) to the category it starts.
SECOND_TABLE = {
    "association ( bikers )": "BIKER_ASSOCIATION",
}

# Cells the marketing team used to mean "we could not find this".
PLACEHOLDERS = {"not publicaly listed", "not publically listed", "not publicly listed", "n/a", "na", "-", ""}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def norm(v) -> str:
    """Cell -> clean string: ints keep no decimal, NBSPs and runs of space collapse."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


def is_placeholder(s: str) -> bool:
    return s.lower().strip() in PLACEHOLDERS


def digits(s: str) -> str:
    return re.sub(r"\D", "", s)


def norm_phone_part(p: str):
    """One phone -> +256XXXXXXXXX when it clearly parses, else the raw text."""
    d = digits(p)
    if len(d) < 7:
        return None                                  # fragment like "343688"
    if d.startswith("2560") and len(d) == 13:        # "256 (0) 312 200 400"
        return "+256" + d[4:]
    if d.startswith("256") and len(d) == 12:
        return "+256" + d[3:]
    if d.startswith("0") and len(d) == 10:
        return "+256" + d[1:]
    if len(d) == 9:                                  # bare "757533759"
        return "+256" + d
    return p.strip()                                 # unusual — keep verbatim, flag upstream


def clean_phone(v: str):
    """Return (phone_or_None, parsed_cleanly)."""
    if not v or is_placeholder(v):
        return None, True
    parts = [p for p in re.split(r"[/,]", v) if digits(p)]
    out, clean = [], True
    for p in parts:
        n = norm_phone_part(p)
        if n is None:
            continue
        if not n.startswith("+256"):
            clean = False
        if n not in out:
            out.append(n)
    return (" / ".join(out) if out else None), clean


def clean_email(v: str):
    """Return (email_or_None, ok). Takes the first of a multi-value cell."""
    if not v or is_placeholder(v):
        return None, True                            # absent, not malformed
    first = re.split(r"[,/;]", v)[0]
    e = first.strip().replace(" ", "").lower().rstrip(".")
    e = e.replace("(at)", "@")                       # "info(at)steelworks…" is unambiguous
    if not e or "example.com" in e:
        return None, False
    return (e, True) if EMAIL_RE.match(e) else (None, False)


def header_row(cells) -> bool:
    joined = " ".join(c.lower() for c in cells)
    return "location" in joined and ("email" in joined or "contact" in joined)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    rows_out, seen, issues = [], {}, []

    for ws in wb.worksheets:
        base_cat = CATEGORY.get(ws.title.strip().lower())
        if not base_cat:
            if ws.max_row > 1:
                issues.append(f"skipped unmapped sheet {ws.title!r}")
            continue

        category = base_cat
        for i, raw in enumerate(ws.iter_rows(values_only=True)):
            cells = [norm(c) for c in raw]
            if not any(cells):
                continue

            first = cells[0].lower().strip()
            if first in SECOND_TABLE:                # a new table starts here
                category = SECOND_TABLE[first]
                continue
            if header_row(cells):
                continue

            name = cells[0]
            if not name:
                issues.append(
                    f"{ws.title} row {i + 1}: no company name "
                    f"(contact: {', '.join(c for c in cells[1:] if c) or 'none'}) — skipped"
                )
                continue

            location = cells[1] if len(cells) > 1 and not is_placeholder(cells[1]) else None
            phone, phone_clean = clean_phone(cells[2] if len(cells) > 2 else "")
            email, email_ok = clean_email(cells[3] if len(cells) > 3 else "")

            flags = []
            if not email_ok:
                flags.append("bad_email")
                issues.append(f"{ws.title} row {i + 1}: {name} — unusable email {cells[3]!r}")
            if not phone_clean:
                flags.append("check_phone")
            if not email and not phone:
                flags.append("no_contact")

            key = (name.lower(), category)
            if key in seen:                          # HOSPITALS / HOSPITALS (2)
                continue
            seen[key] = True

            rows_out.append({
                "name": name,
                "category": category,
                "product": PRODUCT,
                "location": location or None,
                "phone": phone,
                "email": email,
                "outreach_status": "not_contacted",
                "flags": flags or None,
                "source": SOURCE,
            })

    rows_out.sort(key=lambda r: (r["category"], r["name"].lower()))
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(rows_out, f, indent=2, ensure_ascii=False)

    by_cat = {}
    for r in rows_out:
        by_cat[r["category"]] = by_cat.get(r["category"], 0) + 1

    print(f"Wrote {len(rows_out)} SEAL prospects -> {OUT}\n")
    for c, n in sorted(by_cat.items()):
        print(f"  {c:<22} {n:>3}")
    print(f"\n  with email            {sum(1 for r in rows_out if r['email']):>3}")
    print(f"  with phone            {sum(1 for r in rows_out if r['phone']):>3}")
    print(f"  flagged               {sum(1 for r in rows_out if r['flags']):>3}")

    if issues:
        print(f"\nNeeds a human look ({len(issues)}):")
        for m in issues:
            print(f"  - {m}")


if __name__ == "__main__":
    main()
